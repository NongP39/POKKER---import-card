#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Pokker Card Encoder — GUI Desktop Application (standalone .exe)"""

import os
import sys
import json
import math
import struct
import random
import threading
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from PIL import Image, ImageDraw, ImageFont
    _PIL = True
except ImportError:
    _PIL = False

try:
    import openpyxl
    _XLSX = True
except ImportError:
    _XLSX = False

# ── Format constants  (must match Godot decoder) ──────────────────────────────
MAGIC       = b'POKR'
VERSION     = 1
IMG_W       = 512
IMG_H       = 512
HDR_ROWS    = 64
DATA_BYTES  = IMG_W * (IMG_H - HDR_ROWS) * 3
META_SIZE   = 88
MAX_PAYLOAD = DATA_BYTES - META_SIZE

CAT_COLORS = {
    "question": (63, 127, 176), "command": (216, 132, 47),
    "deeptalk": (125, 95, 178), "drink":   (47, 154, 147),
    "minigame": (95, 154, 66),  "secret":  (79, 99, 176),
    "skinship": (207, 95, 147), "special": (201, 150, 47),
    "oops":     (207, 59, 44),
}
BG_COLOR      = (240, 235, 228)
DEFAULT_COLOR = (138, 109, 86)


# ── File readers ──────────────────────────────────────────────────────────────

def read_excel(path: str) -> list:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    col_map: dict = {}
    for i, cell in enumerate(ws[1]):
        if cell.value:
            col_map[str(cell.value).strip().lower()] = i

    def find(*keys):
        for k in keys:
            if k in col_map:
                return col_map[k]
        return None

    ci = {
        "cat":     find("หมวดหมู่", "category", "หมวด"),
        "target":  find("เป้าหมาย", "target",   "ผู้ถูกกระทำ"),
        "content": find("เนื้อหา",  "content",  "คำถาม", "คำสั่ง"),
        "penalty": find("โทษ",      "penalty",  "บทลงโทษ"),
    }

    cards: list = []
    counters: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        def cell(key: str) -> str:
            idx = ci.get(key)
            return str(row[idx]).strip() if (idx is not None and idx < len(row) and row[idx] is not None) else ""

        cat     = cell("cat") or "General"
        content = cell("content")
        if not content or content == "None":
            continue

        p = cat[0].upper()
        counters[p] = counters.get(p, 0) + 1
        cards.append({
            "id":         f"{p}{counters[p]:04d}",
            "Category":   cat.capitalize(),
            "Target":     cell("target"),
            "Content":    content,
            "Penalty":    cell("penalty"),
            "TargetType": "Condition",
        })
    return cards


def read_json(path: str) -> list:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("cards", "data", "items"):
            if isinstance(data[key], list):
                return data[key]
    raise ValueError("JSON ต้องเป็น Array ของการ์ด หรือ {\"cards\": [...]}")


# ── Encoder ───────────────────────────────────────────────────────────────────

def _build_meta(idx: int, total: int, set_id: int,
                off: int, chunk_len: int, full_len: int,
                name: str, creator: str) -> bytes:
    nb   = name.encode("utf-8")[:32].ljust(32, b"\x00")
    cb   = creator.encode("utf-8")[:32].ljust(32, b"\x00")
    meta = (MAGIC
            + struct.pack("<BBBBI", VERSION, idx, total, 0, set_id)
            + struct.pack("<III",   off, chunk_len, full_len)
            + nb + cb)
    assert len(meta) == META_SIZE
    return meta


def _load_font(size: int):
    for path in [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "C:/Windows/Fonts/verdana.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
    return ImageFont.load_default()


def encode_cards(cards: list, name: str, creator: str,
                 out_dir: str, on_progress=None) -> list:
    payload  = json.dumps(cards, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    full_len = len(payload)
    n        = max(1, math.ceil(full_len / MAX_PAYLOAD))
    set_id   = random.randint(0, 0xFFFF_FFFF)

    safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in name).strip() or "pokker_cards"
    hdr  = CAT_COLORS.get((cards[0]["Category"].lower() if cards else ""), DEFAULT_COLOR)
    fnt_lg = _load_font(20)
    fnt_sm = _load_font(13)
    paths  = []

    for idx in range(n):
        off   = idx * MAX_PAYLOAD
        chunk = payload[off: off + MAX_PAYLOAD]
        meta  = _build_meta(idx, n, set_id, off, len(chunk), full_len, name, creator)
        data  = meta + chunk

        buf = bytearray(IMG_W * IMG_H * 3)
        for i in range(0, len(buf), 3):
            buf[i] = BG_COLOR[0]; buf[i+1] = BG_COLOR[1]; buf[i+2] = BG_COLOR[2]
        for r in range(HDR_ROWS):
            for c in range(IMG_W):
                b = (r * IMG_W + c) * 3
                buf[b] = hdr[0]; buf[b+1] = hdr[1]; buf[b+2] = hdr[2]
        ds = HDR_ROWS * IMG_W * 3
        for i, bv in enumerate(data):
            buf[ds + i] = bv

        img  = Image.frombytes("RGB", (IMG_W, IMG_H), bytes(buf))
        draw = ImageDraw.Draw(img)
        disp = name if n == 1 else f"{name}  ({idx+1}/{n})"
        draw.text((12, 8),  disp,              fill=(255, 255, 255), font=fnt_lg)
        if creator:
            draw.text((12, 38), f"by {creator}", fill=(255, 240, 200), font=fnt_sm)

        fname = f"{safe}.png" if n == 1 else f"{safe}_{idx+1}of{n}.png"
        p     = os.path.join(out_dir, fname)
        img.save(p, "PNG", optimize=False)
        paths.append(p)
        if on_progress:
            on_progress(idx + 1, n)

    return paths


# ── GUI ───────────────────────────────────────────────────────────────────────

class App:
    BG    = "#f6f0e8"
    RED   = "#d54a22"
    BROWN = "#8a6d56"
    DARK  = "#3a2c20"
    GREEN = "#4a7c3f"
    FONT  = "Segoe UI"

    def __init__(self, root: tk.Tk):
        self.root    = root
        self.cards   = []
        self.out_dir = None

        root.title("Pokker Card Encoder")
        root.configure(bg=self.BG)
        root.resizable(False, False)

        W, H = 500, 610
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")

        self._build()
        self._check_deps()

    def _f(self, size=11, bold=False):
        return (self.FONT, size, "bold" if bold else "normal")

    def _check_deps(self):
        missing = []
        if not _PIL:  missing.append("Pillow")
        if not _XLSX: missing.append("openpyxl")
        if missing:
            messagebox.showerror("ขาด Library",
                f"กรุณาติดตั้ง:\n  pip install {' '.join(missing)}")

    # ── Build UI ──────────────────────────────────────────────────────────────

    def _build(self):
        # ── Header ──
        hdr = tk.Frame(self.root, bg=self.RED)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="🃏  Pokker Card Encoder",
                 font=self._f(17, True), fg="white", bg=self.RED, pady=12).pack()
        tk.Label(hdr, text="แปลง Excel เป็นภาพสำหรับ import เข้าเกม Pokker",
                 font=self._f(10), fg="#ffd0c0", bg=self.RED).pack(pady=(0, 12))

        body = tk.Frame(self.root, bg=self.BG, padx=22)
        body.pack(fill=tk.BOTH, expand=True, pady=8)

        # ── Step 1: File ──────────────────────────────────────────────────
        self._heading(body, "1", "เลือกไฟล์การ์ด")

        drop = tk.Frame(body, bg="#fff4ed", relief="solid", bd=2, cursor="hand2")
        drop.pack(fill=tk.X, pady=(0, 6))

        self.lbl_drop_icon = tk.Label(drop, text="📂",
                                       font=self._f(32), bg="#fff4ed", cursor="hand2", pady=4)
        self.lbl_drop_icon.pack()
        tk.Label(drop, text="คลิกเพื่อเลือกไฟล์ Excel หรือ JSON",
                 font=self._f(12, True), fg=self.RED, bg="#fff4ed", cursor="hand2").pack()
        tk.Label(drop, text=".xlsx  ·  .xls  ·  .json",
                 font=self._f(10), fg=self.BROWN, bg="#fff4ed", cursor="hand2").pack()
        tk.Label(drop, text="คอลัมน์: ลำดับ  ·  หมวดหมู่  ·  เป้าหมาย  ·  เนื้อหา  ·  โทษ",
                 font=self._f(9), fg="#bbb", bg="#fff4ed", cursor="hand2", pady=8).pack()

        self._bind_click(drop, self._browse)

        self.lbl_file  = tk.Label(body, text="", font=self._f(10),
                                   fg=self.DARK, bg=self.BG, anchor="w")
        self.lbl_file.pack(fill=tk.X)
        self.lbl_count = tk.Label(body, text="", font=self._f(10, True),
                                   fg=self.GREEN, bg=self.BG, anchor="w")
        self.lbl_count.pack(fill=tk.X, pady=(0, 10))

        # ── Step 2: Metadata ──────────────────────────────────────────────
        self._heading(body, "2", "ข้อมูลชุดการ์ด")

        self.var_name    = tk.StringVar()
        self.var_creator = tk.StringVar()
        self._input(body, "ชื่อชุดการ์ด *",          self.var_name)
        self._input(body, "ชื่อผู้สร้าง  (ไม่บังคับ)", self.var_creator)

        # ── Step 3: Encode ────────────────────────────────────────────────
        self._heading(body, "3", "สร้างภาพ PNG")

        self.btn_go = tk.Button(
            body, text="🖼️   สร้างภาพ",
            font=self._f(13, True), bg=self.RED, fg="white",
            activebackground="#b83d1b", activeforeground="white",
            relief="flat", bd=0, pady=12, cursor="hand2",
            state="disabled", command=self._encode)
        self.btn_go.pack(fill=tk.X, pady=(0, 8))

        s = ttk.Style()
        s.theme_use("default")
        s.configure("P.Horizontal.TProgressbar",
                    troughcolor="#e8ddd0", background=self.RED, thickness=8)
        self.prog = ttk.Progressbar(body, style="P.Horizontal.TProgressbar", mode="determinate")
        self.prog.pack(fill=tk.X, pady=(0, 4))

        self.lbl_status = tk.Label(body, text="", font=self._f(10),
                                    fg=self.BROWN, bg=self.BG, anchor="w")
        self.lbl_status.pack(fill=tk.X, pady=(0, 8))

        self.btn_open = tk.Button(
            body, text="📂   เปิดโฟลเดอร์ผลลัพธ์",
            font=self._f(11), bg=self.GREEN, fg="white",
            activebackground="#3a6030", activeforeground="white",
            relief="flat", bd=0, pady=9, cursor="hand2",
            state="disabled", command=self._open_folder)
        self.btn_open.pack(fill=tk.X)

    # ── Widget helpers ────────────────────────────────────────────────────────

    def _heading(self, parent, num, text):
        f = tk.Frame(parent, bg=self.BG)
        f.pack(fill=tk.X, pady=(14, 6))
        tk.Label(f, text=f" {num} ", font=self._f(9, True),
                 bg=self.RED, fg="white", padx=4).pack(side=tk.LEFT)
        tk.Label(f, text=f"  {text}", font=self._f(11, True),
                 bg=self.BG, fg=self.DARK).pack(side=tk.LEFT)

    def _input(self, parent, label, var):
        tk.Label(parent, text=label, font=self._f(10),
                 fg=self.DARK, bg=self.BG, anchor="w").pack(fill=tk.X)
        tk.Entry(parent, textvariable=var, font=self._f(11),
                 bg="white", relief="solid", bd=1, fg=self.DARK).pack(
            fill=tk.X, pady=(2, 10), ipady=7)

    def _bind_click(self, widget, cmd):
        widget.bind("<Button-1>", lambda _: cmd())
        for child in widget.winfo_children():
            child.bind("<Button-1>", lambda _: cmd())

    # ── Actions ───────────────────────────────────────────────────────────────

    def _browse(self):
        path = filedialog.askopenfilename(
            title="เลือกไฟล์การ์ด",
            filetypes=[
                ("ไฟล์การ์ด", "*.xlsx *.xls *.json"),
                ("Excel",     "*.xlsx *.xls"),
                ("JSON",      "*.json"),
                ("ทุกไฟล์",   "*.*"),
            ]
        )
        if path:
            self._load(path)

    def _load(self, path: str):
        try:
            ext = Path(path).suffix.lower()
            if ext in (".xlsx", ".xls"):
                if not _XLSX:
                    messagebox.showerror("ข้อผิดพลาด", "ไม่สามารถอ่าน Excel ได้")
                    return
                self.cards = read_excel(path)
            elif ext == ".json":
                self.cards = read_json(path)
            else:
                messagebox.showerror("ไฟล์ไม่รองรับ", "รองรับเฉพาะ .xlsx, .xls และ .json")
                return

            cats = len(set(c["Category"] for c in self.cards))
            self.lbl_drop_icon.config(text="✅")
            self.lbl_file.config(text=f"📄  {Path(path).name}", fg=self.DARK)
            self.lbl_count.config(text=f"✅  {len(self.cards)} ใบ  ·  {cats} หมวดหมู่")
            if not self.var_name.get():
                self.var_name.set(Path(path).stem)
            self.btn_go.config(state="normal")

        except Exception as e:
            messagebox.showerror("อ่านไฟล์ไม่สำเร็จ", str(e))

    def _encode(self):
        if not self.cards:
            return

        name    = self.var_name.get().strip() or "Pokker Cards"
        creator = self.var_creator.get().strip()

        self.out_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Pokker Export")
        os.makedirs(self.out_dir, exist_ok=True)

        self.btn_go.config(state="disabled", text="⏳  กำลังสร้าง...")
        self.btn_open.config(state="disabled")
        self.prog["value"] = 0
        self.lbl_status.config(text="กำลังประมวลผล...", fg=self.BROWN)

        def worker():
            try:
                def cb(done, total):
                    pct = done / total * 100
                    self.root.after(0, lambda: self.prog.config(value=pct))
                    self.root.after(0, lambda: self.lbl_status.config(
                        text=f"กำลังสร้างภาพที่ {done}/{total}...", fg=self.BROWN))

                paths = encode_cards(self.cards, name, creator, self.out_dir, cb)
                self.root.after(0, lambda: self._done(paths))
            except Exception as e:
                err = str(e)
                self.root.after(0, lambda: self._error(err))

        threading.Thread(target=worker, daemon=True).start()

    def _done(self, paths: list):
        n = len(paths)
        self.btn_go.config(state="normal", text="🖼️   สร้างภาพ")
        self.btn_open.config(state="normal")
        self.prog["value"] = 100
        self.lbl_status.config(
            text=f"✅  สร้างเสร็จ! {n} ภาพ  →  Desktop / Pokker Export",
            fg=self.GREEN)

        if n > 1:
            messagebox.showinfo(
                "สร้างเสร็จแล้ว! 🎉",
                f"สร้าง {n} ภาพเรียบร้อย!\n\n"
                "⚠️  ต้อง import ทุกภาพพร้อมกันในเกม\n"
                "   (เลือกทุกไฟล์พร้อมกันตอน import)\n\n"
                f"บันทึกที่:\n{self.out_dir}")
        else:
            self._open_folder()

    def _error(self, msg: str):
        self.btn_go.config(state="normal", text="🖼️   สร้างภาพ")
        self.lbl_status.config(text="❌  เกิดข้อผิดพลาด กรุณาลองใหม่", fg=self.RED)
        messagebox.showerror("เกิดข้อผิดพลาด", msg)

    def _open_folder(self):
        if self.out_dir and os.path.exists(self.out_dir):
            os.startfile(self.out_dir)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
